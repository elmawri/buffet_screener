"""
Capital Allocation Sheet Populator - V3
FMP for shares change with Yahoo fallback
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_capital_allocation_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['CapitalAllocation']
    cols = COLUMN_MAP['CapitalAllocation']
    
    print(f"\n{'='*80}")
    print(f"POPULATING CAPITAL ALLOCATION SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            info = basic.get('full_info', {})
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Shares change (FMP Phase 3)
            fmp_data = all_data.get('phase3_fmp', {})
            shares_change = fmp_data.get('shares_change', {}).get('shares_5y_change_pct')
            
            ws.cell(row=row, column=cols['Shares_Outstanding_5Y_Change']).value = shares_change
            if shares_change:
                print(f"    Shares 5Y Change: {shares_change:+.1f}% (FMP)")
            
            # Dividend payout (Yahoo)
            dividend_rate = info.get('dividendRate', 0)
            trailing_eps = info.get('trailingEps', 0)
            payout_ratio = None
            if trailing_eps and trailing_eps > 0:
                payout_ratio = (dividend_rate / trailing_eps) * 100
                ws.cell(row=row, column=cols['Dividend_Payout_Ratio']).value = payout_ratio
                print(f"    Payout Ratio: {payout_ratio:.1f}% (Yahoo)")
            
            # Other metrics (placeholders)
            for col in ['Net_Buyback_Yield', 'Reinvestment_Rate', 'Acquisition_Spend_5Y',
                       'Post_MA_ROIC', 'Debt_Issuance_vs_FCF', 'Specials_One_Offs', 'Notes']:
                ws.cell(row=row, column=cols[col]).value = None
            
            # Score
            score_data = {
                'shares_5y_change_pct': shares_change if shares_change else 0,
                'avg_buyback_pe': info.get('trailingPE', 20),
                'dividend_payout_ratio': payout_ratio if payout_ratio else 50,
                'reinvestment_rate': 50,
                'roic': 15
            }
            auto_score = BuffettScorer.calculate_capital_allocation_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            sources = ["Yahoo"]
            if shares_change: sources.append("FMP")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("CAPITAL ALLOCATION SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_capital_allocation_sheet(tickers=sys.argv[1:])
    else:
        populate_capital_allocation_sheet()
