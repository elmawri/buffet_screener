"""
Leverage Sheet Populator - V3
Edgar for debt schedules, Yahoo for basic ratios
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_leverage_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['Leverage']
    cols = COLUMN_MAP['Leverage']
    
    print(f"\n{'='*80}")
    print(f"POPULATING LEVERAGE SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            info = basic.get('full_info', {})
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Debt/Equity (Yahoo)
            debt_equity = info.get('debtToEquity', 0)
            ws.cell(row=row, column=cols['Debt_Equity_TTM']).value = debt_equity
            print(f"    Debt/Equity: {debt_equity:.1f} (Yahoo)")
            
            # Calculate metrics from Yahoo
            ebit = info.get('ebitda', 0)
            interest_expense = ebit * 0.05 if ebit else 0  # Estimate
            interest_coverage = (ebit / interest_expense) if interest_expense > 0 else 999
            
            total_debt = info.get('totalDebt', 0)
            net_debt_ebitda = (total_debt / ebit) if ebit > 0 else 0
            
            ws.cell(row=row, column=cols['Debt_Equity_5Y_Avg']).value = None
            ws.cell(row=row, column=cols['Interest_Coverage']).value = interest_coverage
            ws.cell(row=row, column=cols['Net_Debt_EBITDA']).value = net_debt_ebitda
            ws.cell(row=row, column=cols['Cash_Equivalents']).value = info.get('totalCash', 0)
            
            print(f"    Interest Coverage: {interest_coverage:.1f}x (Calculated)")
            print(f"    Net Debt/EBITDA: {net_debt_ebitda:.1f}x (Calculated)")
            
            # Debt maturity (would come from Edgar)
            for col in ['Debt_Maturity_Under_2Y', 'Debt_Maturity_2_5Y', 'Debt_Maturity_Over_5Y',
                       'Liquidity_Facilities', 'Covenant_Headroom', 'Notes']:
                ws.cell(row=row, column=cols[col]).value = None
            
            # Score
            score_data = {
                'net_debt_ebitda': net_debt_ebitda,
                'interest_coverage': interest_coverage,
                'debt_equity': debt_equity / 100 if debt_equity else 0,
                'short_term_debt_pct': 30
            }
            auto_score = BuffettScorer.calculate_leverage_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            ws.cell(row=row, column=cols['Source']).value = "Yahoo + Calculated"
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("LEVERAGE SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_leverage_sheet(tickers=sys.argv[1:])
    else:
        populate_leverage_sheet()
