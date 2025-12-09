"""
Management Sheet Populator - V3
Uses Edgar for executive tenure, Yahoo for insider ownership
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_management_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['Management']
    cols = COLUMN_MAP['Management']
    
    print(f"\n{'='*80}")
    print(f"POPULATING MANAGEMENT SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            info = basic.get('full_info', {})
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Executive tenure (Edgar Phase 2)
            edgar = all_data.get('phase2_edgar', {})
            execs = edgar.get('executives', {})
            
            ceo_tenure = execs.get('ceo', {}).get('tenure_years')
            cfo_tenure = execs.get('cfo', {}).get('tenure_years')
            
            ws.cell(row=row, column=cols['CEO_Tenure']).value = ceo_tenure
            ws.cell(row=row, column=cols['CFO_Tenure']).value = cfo_tenure
            
            if ceo_tenure:
                print(f"    CEO Tenure: {ceo_tenure} years (Edgar)")
            if cfo_tenure:
                print(f"    CFO Tenure: {cfo_tenure} years (Edgar)")
            
            # Insider Ownership (Yahoo)
            insider_pct = info.get('heldPercentInsiders', 0) * 100
            ws.cell(row=row, column=cols['Insider_Ownership_Pct']).value = insider_pct
            print(f"    Insider Ownership: {insider_pct:.1f}%")
            
            # Other management fields (placeholders)
            for col in ['Compensation_Alignment', 'Capital_Allocation_Letters', 'Disclosure_Quality', 
                       'Related_Party_Risks', 'Accounting_Conservatism', 'Share_Count_5Y_Change',
                       'Buybacks_Below_IV', 'MA_Discipline', 'Dividend_Policy', 'Notes']:
                ws.cell(row=row, column=cols[col]).value = None
            
            # Score
            score_data = {
                'insider_ownership_pct': insider_pct,
                'ceo_tenure_years': ceo_tenure if ceo_tenure else 5,
                'shares_5y_change_pct': 0,
                'avg_pe_ratio': info.get('trailingPE', 20),
                'compensation_aligned': True,
                'dividend_payout_ratio': 50
            }
            auto_score = BuffettScorer.calculate_management_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            sources = ["Yahoo"]
            if ceo_tenure: sources.append("Edgar")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("MANAGEMENT SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_management_sheet(tickers=sys.argv[1:])
    else:
        populate_management_sheet()
