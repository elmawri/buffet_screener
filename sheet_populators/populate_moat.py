"""
Moat Sheet Populator - V3
Uses AI for moat type/pricing power, FMP for historical metrics with Yahoo fallback
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_moat_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['Moat']
    cols = COLUMN_MAP['Moat']
    
    print(f"\n{'='*80}")
    print(f"POPULATING MOAT SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Moat Type (AI Phase 6)
            ai = all_data.get('phase6_ai', {})
            moat_type = ai.get('moat_type')
            ws.cell(row=row, column=cols['Moat_Type']).value = moat_type
            if moat_type:
                print(f"    Moat: {moat_type}")
            
            # Historical metrics (FMP Phase 3, Yahoo Phase 5 fallback)
            historical = coordinator.get_historical_roe_roic()
            fmp_metrics = all_data.get('phase3_fmp', {}).get('metrics_10y', {})
            margins = fmp_metrics.get('margins', {})
            yahoo_fb = all_data.get('phase5_yahoo_fallback', {})
            
            roe_5y = historical.get('roe_5y_avg')
            roic_5y = historical.get('roic_5y_avg')
            gross_5y = margins.get('gross_avg_5y') or yahoo_fb.get('gross_margin_5y')
            op_5y = margins.get('operating_avg_5y') or yahoo_fb.get('operating_margin_5y')
            
            ws.cell(row=row, column=cols['ROE_5Y_Avg']).value = roe_5y
            ws.cell(row=row, column=cols['ROIC_5Y_Avg']).value = roic_5y
            ws.cell(row=row, column=cols['Gross_Margin_5Y_Avg']).value = gross_5y
            ws.cell(row=row, column=cols['Operating_Margin_5Y_Avg']).value = op_5y
            
            if roe_5y:
                print(f"    ROE 5Y: {roe_5y:.1f}% ({historical.get('source', 'Unknown')})")
            if gross_5y:
                source = "FMP" if margins.get('gross_avg_5y') else "Yahoo Fallback"
                print(f"    Gross Margin 5Y: {gross_5y:.1f}% ({source})")
            
            # Market share (placeholder)
            ws.cell(row=row, column=cols['Market_Share_Pct']).value = None
            ws.cell(row=row, column=cols['Market_Share_Trend']).value = None
            
            # Pricing Power (AI Phase 6)
            pricing_power = ai.get('pricing_power')
            ws.cell(row=row, column=cols['Pricing_Power']).value = pricing_power
            if pricing_power:
                print(f"    Pricing Power: {pricing_power}")
            
            # Other moat factors (placeholders)
            for col in ['Customer_Concentration', 'Customer_Switching_Costs', 'Brand_Value', 'Network_Effects', 'Regulatory_Moat', 'Notes']:
                ws.cell(row=row, column=cols[col]).value = None
            
            # Score
            score_data = {
                'moat_type': moat_type if moat_type else 'None',
                'roe_5y_avg': roe_5y if roe_5y else 10,
                'roic_5y_avg': roic_5y if roic_5y else 10,
                'roe_consistency': historical.get('roe_std', 5),
                'gross_margin_5y': gross_5y if gross_5y else 30,
                'pricing_power': pricing_power if pricing_power else 'Low'
            }
            auto_score = BuffettScorer.calculate_moat_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            # Source
            sources = ["Yahoo"]
            if ai: sources.append("AI")
            if fmp_metrics: sources.append("FMP")
            if yahoo_fb: sources.append("Yahoo Fallback")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("MOAT SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_moat_sheet(tickers=sys.argv[1:])
    else:
        populate_moat_sheet()
