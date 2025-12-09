"""
Operating History Sheet Populator - V3
Uses Edgar for founding date, FMP for growth metrics with Yahoo fallback
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_operating_history_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['OperatingHistory']
    cols = COLUMN_MAP['OperatingHistory']
    
    print(f"\n{'='*80}")
    print(f"POPULATING OPERATING HISTORY SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(
                ticker=ticker,
                anthropic_key=ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None,
                fmp_key=FMP_API_KEY,
                fred_key=FRED_API_KEY
            )
            
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            
            # Col 1-2: Ticker, Company
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Col 3-5: Company history (Edgar Phase 2, FMP Phase 3)
            edgar_data = all_data.get('phase2_edgar', {})
            fmp_data = all_data.get('phase3_fmp', {})
            
            founded_year = edgar_data.get('history', {}).get('founded_year')
            ws.cell(row=row, column=cols['Founded_Year']).value = founded_year
            if founded_year:
                print(f"    Founded: {founded_year} (Edgar)")
            
            ipo_year = fmp_data.get('profile', {}).get('ipo_date')
            if ipo_year:
                ipo_year = int(ipo_year.split('-')[0])
                ws.cell(row=row, column=cols['IPO_Year']).value = ipo_year
                years_since_ipo = datetime.now().year - ipo_year
                ws.cell(row=row, column=cols['Years_Since_IPO']).value = years_since_ipo
                print(f"    IPO: {ipo_year}, {years_since_ipo} years (FMP)")
            
            # Col 6: Years Profitable (placeholder)
            ws.cell(row=row, column=cols['Years_Profitability']).value = None
            
            # Col 7-8: Growth metrics (FMP Phase 3 primary, Yahoo Phase 5 fallback)
            metrics_10y = fmp_data.get('metrics_10y', {})
            growth = metrics_10y.get('growth', {})
            yahoo_fallback = all_data.get('phase5_yahoo_fallback', {})
            
            revenue_cagr = growth.get('revenue_cagr_10y') or yahoo_fallback.get('revenue_growth')
            eps_cagr = growth.get('eps_cagr_10y') or yahoo_fallback.get('earnings_growth')
            
            ws.cell(row=row, column=cols['Revenue_CAGR']).value = revenue_cagr
            ws.cell(row=row, column=cols['EPS_CAGR']).value = eps_cagr
            
            if revenue_cagr:
                source = "FMP 10Y" if growth.get('revenue_cagr_10y') else "Yahoo TTM"
                print(f"    Revenue CAGR: {revenue_cagr:.1f}% ({source})")
            if eps_cagr:
                source = "FMP 10Y" if growth.get('eps_cagr_10y') else "Yahoo TTM"
                print(f"    EPS CAGR: {eps_cagr:.1f}% ({source})")
            
            # Col 9-10: Down years (placeholder - would need detailed analysis)
            ws.cell(row=row, column=cols['Rev_Down_Years']).value = None
            ws.cell(row=row, column=cols['EPS_Down_Years']).value = None
            
            # Col 11-13: Restatements and pivots (Edgar Phase 2)
            restatements = edgar_data.get('restatements', [])
            ws.cell(row=row, column=cols['Restatements']).value = len(restatements) if restatements else 0
            ws.cell(row=row, column=cols['Major_Pivots']).value = None
            ws.cell(row=row, column=cols['Notes']).value = None
            
            # Col 14: Score
            score_data = {
                'years_profitable': 10,  # Default
                'revenue_cagr': revenue_cagr if revenue_cagr else 5,
                'eps_cagr': eps_cagr if eps_cagr else 5,
                'restatements': len(restatements) if restatements else 0
            }
            auto_score = BuffettScorer.calculate_operating_history_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            # Col 15: Source
            sources = ["Yahoo"]
            if founded_year:
                sources.append("Edgar")
            if growth.get('revenue_cagr_10y'):
                sources.append("FMP")
            elif yahoo_fallback.get('revenue_growth'):
                sources.append("Yahoo Fallback")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            
            # Col 16: Last Updated
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("OPERATING HISTORY SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_operating_history_sheet(tickers=sys.argv[1:])
    else:
        populate_operating_history_sheet()
