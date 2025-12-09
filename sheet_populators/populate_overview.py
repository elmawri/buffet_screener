"""
Overview Sheet Populator - V3
Fixes formulas to reference correct columns in all source sheets
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from config import TICKERS, EXCEL_FILE
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP

def populate_overview_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws_overview = wb['Overview']
    
    cols_overview = COLUMN_MAP['Overview']
    cols_tickers = COLUMN_MAP['Tickers']
    
    # Score column positions in each sheet
    score_columns = {
        'Simplicity': COLUMN_MAP['Simplicity']['Score'],
        'OperatingHistory': COLUMN_MAP['OperatingHistory']['Score'],
        'Moat': COLUMN_MAP['Moat']['Score'],
        'Management': COLUMN_MAP['Management']['Score'],
        'ROE_ROIC': COLUMN_MAP['ROE_ROIC']['Score'],
        'Predictability': COLUMN_MAP['Predictability']['Score'],
        'CapitalAllocation': COLUMN_MAP['CapitalAllocation']['Score'],
        'Leverage': COLUMN_MAP['Leverage']['Score'],
        'Resilience': COLUMN_MAP['Resilience']['Score'],
        'PriceValue': COLUMN_MAP['PriceValue']['Score']
    }
    
    print(f"\n{'='*80}")
    print(f"POPULATING OVERVIEW SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for idx, ticker in enumerate(tickers, start=2):
        print(f"  Setting up {ticker} (row {row})...")
        
        # Col A-B: Ticker, Company
        ws_overview.cell(row=row, column=cols_overview['Ticker']).value = f"=Tickers!A{idx}"
        ws_overview.cell(row=row, column=cols_overview['Company']).value = f"=Tickers!B{idx}"
        
        # Col C-E: Sector, Industry, Country (CORRECTED columns)
        sector_col = get_column_letter(cols_tickers['Sector'])
        industry_col = get_column_letter(cols_tickers['Industry'])
        country_col = get_column_letter(cols_tickers['Country'])
        
        ws_overview.cell(row=row, column=cols_overview['Sector']).value = f"=Tickers!{sector_col}{idx}"
        ws_overview.cell(row=row, column=cols_overview['Industry']).value = f"=Tickers!{industry_col}{idx}"
        ws_overview.cell(row=row, column=cols_overview['Country']).value = f"=Tickers!{country_col}{idx}"
        
        # Col F: Notes (empty)
        
        # Cols G-P: Pull scores from CORRECT columns
        ws_overview.cell(row=row, column=cols_overview['Simplicity_Score']).value = f"=Simplicity!${get_column_letter(score_columns['Simplicity'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['OperatingHistory_Score']).value = f"=OperatingHistory!${get_column_letter(score_columns['OperatingHistory'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['Moat_Score']).value = f"=Moat!${get_column_letter(score_columns['Moat'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['Management_Score']).value = f"=Management!${get_column_letter(score_columns['Management'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['ROE_ROIC_Score']).value = f"=ROE_ROIC!${get_column_letter(score_columns['ROE_ROIC'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['Predictability_Score']).value = f"=Predictability!${get_column_letter(score_columns['Predictability'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['CapitalAllocation_Score']).value = f"=CapitalAllocation!${get_column_letter(score_columns['CapitalAllocation'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['Leverage_Score']).value = f"=Leverage!${get_column_letter(score_columns['Leverage'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['Resilience_Score']).value = f"=Resilience!${get_column_letter(score_columns['Resilience'])}${idx}"
        ws_overview.cell(row=row, column=cols_overview['PriceValue_Score']).value = f"=PriceValue!{get_column_letter(score_columns['PriceValue'])}{idx}"
        
        # Col Q: Total
        ws_overview.cell(row=row, column=cols_overview['Total']).value = f"=SUM(G{row}:P{row})"
        
        # Col R: Average
        ws_overview.cell(row=row, column=cols_overview['Average']).value = f"=Q{row}/10"
        
        # Col S: Rank
        total_rows = len(tickers) + 1
        ws_overview.cell(row=row, column=cols_overview['Rank']).value = f"=RANK(R{row},R$2:R${total_rows},0)"
        
        row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("OVERVIEW SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_overview_sheet(tickers=sys.argv[1:])
    else:
        populate_overview_sheet()
