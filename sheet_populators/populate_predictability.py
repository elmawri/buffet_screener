"""
Predictability Sheet Populator - V3
FMP for volatility metrics with Yahoo fallback
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_predictability_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['Predictability']
    cols = COLUMN_MAP['Predictability']
    
    print(f"\n{'='*80}")
    print(f"POPULATING PREDICTABILITY SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Volatility metrics (FMP Phase 3)
            fmp_metrics = all_data.get('phase3_fmp', {}).get('metrics_10y', {})
            volatility = fmp_metrics.get('volatility', {})
            margins = fmp_metrics.get('margins', {})
            
            # Revenue/EPS volatility
            rev_std = volatility.get('revenue_std')
            rev_cov = volatility.get('revenue_cov')
            
            ws.cell(row=row, column=cols['EPS_10Y_StdDev']).value = None  # Would need from FMP
            ws.cell(row=row, column=cols['EPS_10Y_CAGR']).value = fmp_metrics.get('growth', {}).get('eps_cagr_10y')
            ws.cell(row=row, column=cols['Operating_Margin_10Y_Avg']).value = margins.get('operating_avg_5y')
            ws.cell(row=row, column=cols['Operating_Margin_10Y_StdDev']).value = margins.get('operating_std_10y')
            ws.cell(row=row, column=cols['Revenue_10Y_StdDev']).value = rev_std
            
            if rev_std:
                print(f"    Revenue StdDev: {rev_std:.1f}% (FMP)")
            if margins.get('operating_std_10y'):
                print(f"    Op Margin StdDev: {margins['operating_std_10y']:.1f}% (FMP)")
            
            # Margin stability
            gross_std = margins.get('gross_std_10y', 5)
            if gross_std < 2:
                stability = "High"
            elif gross_std < 5:
                stability = "Med"
            else:
                stability = "Low"
            ws.cell(row=row, column=cols['Gross_Margin_Stability']).value = stability
            
            # Other metrics (placeholders)
            ws.cell(row=row, column=cols['Earnings_Drawdown']).value = None
            ws.cell(row=row, column=cols['Guidance_Accuracy']).value = None
            ws.cell(row=row, column=cols['Notes']).value = None
            
            # Score
            score_data = {
                'revenue_cov': rev_cov if rev_cov else 15,
                'earnings_cov': 15,  # Default
                'margin_stability': margins.get('operating_std_10y', 5)
            }
            auto_score = BuffettScorer.calculate_predictability_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            sources = ["Yahoo"]
            if fmp_metrics: sources.append("FMP")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("PREDICTABILITY SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_predictability_sheet(tickers=sys.argv[1:])
    else:
        populate_predictability_sheet()
