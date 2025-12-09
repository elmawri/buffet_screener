"""
Price/Value Sheet Populator - V3
FMP for historical medians with Yahoo fallback, FRED for treasury
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_price_value_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['PriceValue']
    cols = COLUMN_MAP['PriceValue']
    
    print(f"\n{'='*80}")
    print(f"POPULATING PRICE/VALUE SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            info = basic.get('full_info', {})
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Current valuation (Yahoo)
            pe_ttm = info.get('trailingPE')
            pb_mrq = info.get('priceToBook')
            ev_ebit_ttm = info.get('enterpriseToEbitda')
            
            ws.cell(row=row, column=cols['PE_TTM']).value = pe_ttm
            ws.cell(row=row, column=cols['PB_MRQ']).value = pb_mrq
            ws.cell(row=row, column=cols['EV_EBIT_TTM']).value = ev_ebit_ttm
            
            print(f"    P/E TTM: {pe_ttm:.1f} (Yahoo)" if pe_ttm else "    P/E TTM: N/A")
            
            # Historical medians (FMP Phase 3, Yahoo Phase 5 fallback)
            fmp_metrics = all_data.get('phase3_fmp', {}).get('metrics_10y', {})
            valuation = fmp_metrics.get('valuation', {})
            yahoo_fb = all_data.get('phase5_yahoo_fallback', {})
            
            pe_median = valuation.get('pe_median_10y') or yahoo_fb.get('pe_median')
            pb_median = valuation.get('pb_median_10y') or yahoo_fb.get('pb_median')
            ev_ebit_median = valuation.get('ev_ebit_median_10y')
            
            ws.cell(row=row, column=cols['PE_10Y_Median']).value = pe_median
            ws.cell(row=row, column=cols['PB_10Y_Median']).value = pb_median
            ws.cell(row=row, column=cols['EV_EBIT_10Y_Median']).value = ev_ebit_median
            
            if pe_median:
                source = "FMP" if valuation.get('pe_median_10y') else "Yahoo Fallback"
                print(f"    P/E Median: {pe_median:.1f} ({source})")
            
            # P/FCF and other metrics (placeholders)
            ws.cell(row=row, column=cols['P_FCF_TTM']).value = None
            ws.cell(row=row, column=cols['P_FCF_10Y_Median']).value = None
            
            # Treasury yield (FRED Phase 4)
            treasury = all_data.get('phase4_fred', {}).get('treasury_10y')
            
            # Yield spread
            if treasury and pe_ttm and pe_ttm > 0:
                earnings_yield = (1 / pe_ttm) * 100
                spread_bps = (earnings_yield - treasury) * 100
                ws.cell(row=row, column=cols['Yield_vs_10Y_UST']).value = spread_bps
                print(f"    Yield Spread: {spread_bps:.0f} bps (vs {treasury:.2f}% Treasury)")
            
            # Other fields (placeholders)
            ws.cell(row=row, column=cols['Implied_IV_Discount']).value = None
            ws.cell(row=row, column=cols['Alt_Fair_Value']).value = None
            ws.cell(row=row, column=cols['Notes']).value = None
            
            # Score (Manual - leave empty)
            ws.cell(row=row, column=cols['Score']).value = None
            
            # Source
            sources = ["Yahoo"]
            if fmp_metrics: sources.append("FMP")
            if yahoo_fb: sources.append("Yahoo Fallback")
            if treasury: sources.append("FRED")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("PRICE/VALUE SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_price_value_sheet(tickers=sys.argv[1:])
    else:
        populate_price_value_sheet()
