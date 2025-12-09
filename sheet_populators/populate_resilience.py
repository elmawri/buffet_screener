"""
Resilience Sheet Populator - V3
FMP for crisis performance, AI for demand type, Yahoo for dividend cuts
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3
import yfinance as yf

def populate_resilience_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['Resilience']
    cols = COLUMN_MAP['Resilience']
    
    print(f"\n{'='*80}")
    print(f"POPULATING RESILIENCE SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            info = basic.get('full_info', {})
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # Beta (Yahoo)
            beta = info.get('beta', 1.0)
            ws.cell(row=row, column=cols['Beta_5Y']).value = beta
            print(f"    Beta: {beta:.2f} (Yahoo)")
            
            # Demand Type (AI Phase 6)
            ai = all_data.get('phase6_ai', {})
            demand_type = ai.get('demand_type')
            ws.cell(row=row, column=cols['Demand_Type']).value = demand_type
            if demand_type:
                print(f"    Demand Type: {demand_type} (AI)")
            
            # Op Margin StdDev (FMP)
            fmp_metrics = all_data.get('phase3_fmp', {}).get('metrics_10y', {})
            margins = fmp_metrics.get('margins', {})
            op_margin_std = margins.get('operating_std_10y')
            ws.cell(row=row, column=cols['Op_Margin_10Y_StdDev']).value = op_margin_std
            
            # Crisis performance (FMP Phase 3)
            crisis = all_data.get('phase3_fmp', {}).get('crisis_performance', {})
            
            rev_2008 = crisis.get('2008_2009', {}).get('revenue_change_pct')
            eps_2008 = crisis.get('2008_2009', {}).get('eps_change_pct')
            rev_2020 = crisis.get('2020', {}).get('revenue_change_pct')
            eps_2020 = crisis.get('2020', {}).get('eps_change_pct')
            
            ws.cell(row=row, column=cols['Revenue_Change_2008_09']).value = rev_2008
            ws.cell(row=row, column=cols['EPS_Change_2008_09']).value = eps_2008
            ws.cell(row=row, column=cols['Revenue_Change_2020']).value = rev_2020
            ws.cell(row=row, column=cols['EPS_Change_2020']).value = eps_2020
            
            if rev_2008:
                print(f"    2008-09: Rev {rev_2008:+.1f}%, EPS {eps_2008:+.1f}% (FMP)")
            if rev_2020:
                print(f"    2020: Rev {rev_2020:+.1f}%, EPS {eps_2020:+.1f}% (FMP)")
            
            # Dividend cuts (Yahoo)
            stock = yf.Ticker(ticker)
            dividend_cuts = False
            try:
                dividends = stock.dividends
                if not dividends.empty and len(dividends) > 4:
                    for i in range(1, min(len(dividends), 20)):
                        if dividends.iloc[i] < dividends.iloc[i-1] * 0.9:
                            dividend_cuts = True
                            break
            except:
                pass
            ws.cell(row=row, column=cols['Dividend_Cuts_Crises']).value = "Yes" if dividend_cuts else "No"
            
            # Other metrics (placeholders)
            for col in ['Customer_Diversification', 'Supply_Chain_Risk', 'Regulatory_Sensitivity', 'Notes']:
                ws.cell(row=row, column=cols[col]).value = None
            
            # Score
            score_data = {
                'revenue_change_2008': rev_2008 if rev_2008 else 0,
                'revenue_change_2020': rev_2020 if rev_2020 else 0,
                'dividend_cuts': dividend_cuts,
                'demand_type': demand_type if demand_type else 'Mixed',
                'customer_diversification': 'Medium'
            }
            auto_score = BuffettScorer.calculate_resilience_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            sources = ["Yahoo"]
            if crisis.get('2008_2009'): sources.append("FMP")
            if ai: sources.append("AI")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("RESILIENCE SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_resilience_sheet(tickers=sys.argv[1:])
    else:
        populate_resilience_sheet()
