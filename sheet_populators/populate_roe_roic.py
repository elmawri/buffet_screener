"""
ROE/ROIC Sheet Populator - V3
FMP for historical averages with Yahoo fallback
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_roe_roic_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['ROE_ROIC']
    cols = COLUMN_MAP['ROE_ROIC']
    
    print(f"\n{'='*80}")
    print(f"POPULATING ROE/ROIC SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(ticker, ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None, FMP_API_KEY, FRED_API_KEY)
            all_data = coordinator.get_all_data()
            basic = coordinator.get_basic_info()
            info = basic.get('full_info', {})
            
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            print(f"\n  {ticker}:")
            
            # ROE TTM (Yahoo)
            roe_ttm = info.get('returnOnEquity', 0) * 100 if info.get('returnOnEquity') else 0
            ws.cell(row=row, column=cols['ROE_TTM']).value = roe_ttm
            print(f"    ROE TTM: {roe_ttm:.1f}% (Yahoo)")
            
            # Historical ROE/ROIC (FMP Phase 3, Yahoo Phase 5 fallback)
            historical = coordinator.get_historical_roe_roic()
            
            roe_5y = historical.get('roe_5y_avg')
            roe_10y = historical.get('roe_10y_avg')
            roic_ttm = roe_ttm * 0.7  # Estimate
            roic_5y = historical.get('roic_5y_avg')
            roic_10y = historical.get('roic_10y_avg')
            
            ws.cell(row=row, column=cols['ROE_5Y_Avg']).value = roe_5y
            ws.cell(row=row, column=cols['ROIC_TTM']).value = roic_ttm
            ws.cell(row=row, column=cols['ROIC_5Y_Avg']).value = roic_5y
            
            if roe_5y:
                print(f"    ROE 5Y Avg: {roe_5y:.1f}% ({historical.get('source', 'Unknown')})")
            if roic_5y:
                print(f"    ROIC 5Y Avg: {roic_5y:.1f}% ({historical.get('source', 'Unknown')})")
            
            # Other metrics (placeholders)
            ws.cell(row=row, column=cols['FCF_Margin_5Y_Avg']).value = None
            ws.cell(row=row, column=cols['CapEx_Sales_5Y_Avg']).value = None
            ws.cell(row=row, column=cols['Notes']).value = None
            
            # Score
            score_data = {
                'roe_ttm': roe_ttm,
                'roe_5y_avg': roe_5y if roe_5y else roe_ttm,
                'roic_ttm': roic_ttm,
                'roic_5y_avg': roic_5y if roic_5y else roic_ttm,
                'roe_consistency': historical.get('roe_std', 5)
            }
            auto_score = BuffettScorer.calculate_roe_roic_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Score: {auto_score}/10")
            
            sources = ["Yahoo"]
            if all_data.get('phase3_fmp', {}).get('metrics_10y'):
                sources.append("FMP")
            if all_data.get('phase5_yahoo_fallback'):
                sources.append("Yahoo Fallback")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("ROE/ROIC SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_roe_roic_sheet(tickers=sys.argv[1:])
    else:
        populate_roe_roic_sheet()
