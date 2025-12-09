"""
Simplicity Sheet Populator - V3 with DataCoordinatorV3
Uses Edgar for segments, AI for qualitative analysis
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from scoring.scoring_engine import BuffettScorer
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_simplicity_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['Simplicity']
    cols = COLUMN_MAP['Simplicity']
    
    print(f"\n{'='*80}")
    print(f"POPULATING SIMPLICITY SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        try:
            coordinator = DataCoordinatorV3(
                ticker=ticker,
                anthropic_key=ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None,
                fmp_key=FMP_API_KEY,
                fred_key=FRED_API_KEY
            )
            
            # Get ALL data (6 phases)
            all_data = coordinator.get_all_data()
            
            basic = coordinator.get_basic_info()
            
            # Col 1-2: Ticker, Company
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            # Col 3-4: Segment data (from Edgar Phase 2)
            segments = all_data.get('phase2_edgar', {}).get('segments', {})
            segment_count = segments.get('segment_count', 0)
            segments_list = segments.get('segments', [])
            
            ws.cell(row=row, column=cols['Segment_Count']).value = segment_count if segment_count else None
            ws.cell(row=row, column=cols['Segments_List']).value = ', '.join(segments_list[:5]) if segments_list else None
            
            print(f"\n  {ticker}:")
            print(f"    Segments: {segment_count if segment_count else 'Edgar unavailable'}")
            
            # Col 5-7: Revenue/Geographic data (placeholders - would need detailed Edgar parsing)
            ws.cell(row=row, column=cols['Rev_by_Top_Segment']).value = None
            ws.cell(row=row, column=cols['Geographic_Count']).value = None
            ws.cell(row=row, column=cols['Rev_by_Geography']).value = None
            
            # Col 8: Customer Concentration (placeholder)
            ws.cell(row=row, column=cols['Customer_Concentration']).value = None
            
            # Col 9: Business Model (from AI Phase 6)
            ai_analysis = all_data.get('phase6_ai', {})
            business_model = ai_analysis.get('business_model')
            ws.cell(row=row, column=cols['Business_Model']).value = business_model
            if business_model:
                print(f"    Business Model: {business_model[:60]}...")
            
            # Col 10-17: Other operational metrics (placeholders)
            ws.cell(row=row, column=cols['Products_SKUs']).value = None
            ws.cell(row=row, column=cols['Manufacturing']).value = None
            ws.cell(row=row, column=cols['Distribution']).value = None
            ws.cell(row=row, column=cols['Vertical_Integration']).value = None
            ws.cell(row=row, column=cols['Regulatory_Complexity']).value = None
            ws.cell(row=row, column=cols['Technology_Complexity']).value = None
            ws.cell(row=row, column=cols['Restatements']).value = None
            ws.cell(row=row, column=cols['Major_Pivots']).value = None
            
            # Col 18: Score (uses complexity from AI Phase 6)
            complexity_score = ai_analysis.get('complexity_score', 5)
            
            score_data = {
                'segment_count': segment_count if segment_count else 5,
                'geographic_diversity': 0,  # Placeholder
                'complexity_score': complexity_score
            }
            auto_score = BuffettScorer.calculate_simplicity_score(score_data)
            ws.cell(row=row, column=cols['Score']).value = auto_score
            print(f"    Complexity: {complexity_score}, Score: {auto_score}/10")
            
            # Col 19: Source
            sources = ["Yahoo"]
            if segments.get('segment_count'):
                sources.append("Edgar")
            if ai_analysis:
                sources.append("AI")
            ws.cell(row=row, column=cols['Source']).value = " + ".join(sources)
            
            # Col 20: Last Updated
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"\n  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("SIMPLICITY SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_simplicity_sheet(tickers=sys.argv[1:])
    else:
        populate_simplicity_sheet()
