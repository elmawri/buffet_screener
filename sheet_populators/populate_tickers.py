"""
Tickers Sheet Populator - V3 with DataCoordinatorV3
Populates basic company information with ISIN and CIK
"""
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import load_workbook
from config import TICKERS, EXCEL_FILE, ANTHROPIC_API_KEY, FMP_API_KEY, FRED_API_KEY, USE_AI_ANALYSIS
from datetime import datetime
sys.path.append('.')
from sheet_populators.column_mappings import COLUMN_MAP
from data_fetchers.data_coordinator_v3 import DataCoordinatorV3

def populate_tickers_sheet(tickers=None, excel_file=None):
    if tickers is None: tickers = TICKERS
    if excel_file is None: excel_file = EXCEL_FILE
    
    wb = load_workbook(excel_file)
    ws = wb['Tickers']
    cols = COLUMN_MAP['Tickers']
    
    print(f"\n{'='*80}")
    print(f"POPULATING TICKERS SHEET - V3")
    print(f"{'='*80}\n")
    
    row = 2
    for ticker in tickers:
        print(f"\nProcessing {ticker}...")
        
        try:
            # Initialize coordinator
            coordinator = DataCoordinatorV3(
                ticker=ticker,
                anthropic_key=ANTHROPIC_API_KEY if USE_AI_ANALYSIS else None,
                fmp_key=FMP_API_KEY,
                fred_key=FRED_API_KEY
            )
            
            # Phase 1 only (basic info sufficient for this sheet)
            basic = coordinator.get_basic_info()
            
            # Col 1-2: Ticker, Company
            ws.cell(row=row, column=cols['Ticker']).value = ticker
            ws.cell(row=row, column=cols['Company']).value = basic.get('company_name')
            
            # Col 3-4: ISIN, CIK (THE KEY FIELDS!)
            isin = basic.get('isin')
            cik = basic.get('cik')
            ws.cell(row=row, column=cols['ISIN']).value = isin
            ws.cell(row=row, column=cols['CIK']).value = cik
            
            print(f"  ✅ ISIN: {isin if isin else '❌ Not found'}")
            print(f"  ✅ CIK: {cik if cik else '❌ Not found'}")
            
            # Col 5: Exchange
            exchange = basic.get('full_info', {}).get('exchange')
            ws.cell(row=row, column=cols['Exchange']).value = exchange
            
            # Col 6-7: Sector, Industry
            ws.cell(row=row, column=cols['Sector']).value = basic.get('sector')
            ws.cell(row=row, column=cols['Industry']).value = basic.get('industry')
            
            # Col 8-9: Country, Currency
            ws.cell(row=row, column=cols['Country']).value = basic.get('country')
            ws.cell(row=row, column=cols['Currency']).value = basic.get('full_info', {}).get('currency', 'USD')
            
            # Col 10: Last Updated
            ws.cell(row=row, column=cols['Last_Updated']).value = datetime.now().strftime("%Y-%m-%d")
            
            row += 1
            
        except Exception as e:
            print(f"  ❌ ERROR: {e}")
            import traceback
            traceback.print_exc()
            row += 1
    
    wb.save(excel_file)
    print(f"\n{'='*80}")
    print("TICKERS SHEET UPDATED!")
    print(f"{'='*80}\n")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        populate_tickers_sheet(tickers=sys.argv[1:])
    else:
        populate_tickers_sheet()
